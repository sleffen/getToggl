import base64
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from help_functions import *
import math

wb = load_workbook('mamutliste.xlsx')
ws = wb.active

# Fyll inn usertoken
usertoken = '08ef0fc9b39e62ffd303593b684ce1a2'
# workspace_id='1813882'

today_date = time.strftime("%Y-%m-%d")
#today_date = '2017-01-12'
print('Dato: ' + today_date)

#lager authorization header
string = usertoken + ':api_token'
string = b'08ef0fc9b39e62ffd303593b684ce1a2:api_token'
headers = {'Authorization': 'Basic ' + base64.b64encode(string)}

#forespor apiet for a se om vi har kontakt
url = 'https://www.toggl.com/api/v8/me'
response = requests.get(url, headers=headers)
if response.status_code != 200:
    print ('Login failed. Check your API key')
    quit()

#finner navn, workspace token id og epost adresse
response = response.json()
#pprint.pprint((response))
fornavn = response['data']['fullname'].split()[1]
print('Navn: ' + fornavn)
workspace_id = response['data']['workspaces'][0]['id']
print('WorkspaceID: ' + str(workspace_id))
email = response['data']['email']
print('Email: ' + email)

#Henter ut prosjektliste
params = {'user_agent': email, 'workspace_id': workspace_id}
url = 'https://www.toggl.com/api/v8/workspaces/' + str(workspace_id) + '/projects'
project_list = requests.get(url, headers=headers, params=params).json()

#modifiserer liste til a inneholde flere punkter
project_list_modified = []
for project in project_list:
    project_list_modified.append({'project_id': project['id'], 'project_name': project['name'],
                          'hours': []})


#henter data fra api
day = get_data(today_date, headers, params)
data = day['data']
print('---------------')
#legger inn tider i de prosjektene som ligger med oppforing
for day_data in data:
    for item in project_list_modified:
        item['hours'].append(project_hours(item['project_id'], day_data))

#summerer timer og skriver de ut til console
print ('Prosjekter med timer: ')
total_sum = 0
total_accurate_sum = 0
for item in project_list_modified:
    sum_hours = 0
    for count in item['hours']:
        sum_hours += float(count)
    total_accurate_sum += sum_hours
    sum_hours = (math.ceil(sum_hours*4))/4
    #sum_hours = 0.5*math.ceil(2*sum_hours)
    total_sum += sum_hours

    item['hours']=sum_hours
    if str(item['hours']) != '0.0':
        fPrintToExcel(ws,time.strftime("%d-%m-%Y"),str(item['hours']),item['project_name'])
        print (str(project_list_modified.index(item)) + ': ' + item['project_name'] + " : " + str(item['hours']))

print('\nTotal sum for dagen: ' + str(total_sum) + '/' + str(total_accurate_sum) + ' (' + str(total_sum-total_accurate_sum) + ')')
max_rows = ws.max_row
#cellrange='A'+str(max_rows+1)+':C'+str(max_rows+1)

ws.cell(column=1,row=max_rows+1, value='Total')
ws.cell(column=2,row=max_rows+1, value=total_sum)
wb.save('mamutliste.xlsx')
# clients_or_projects=raw_input('\nGroup by clients or by projects? (p or c) ')
