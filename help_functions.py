import requests
import pprint
import time
import openpyxl

# search the JSON response for occurences of project_id, return project summary time in minutes
def project_hours(project_id,response):

    if response['pid']==project_id:
        return round(response['dur']/float(3600000),2)

    return 0

# get weekly summary starting from start_date
def get_data(start_date,headers,params):
    url='https://toggl.com/reports/api/v2/details'
    params['since']=start_date
    params['until']=start_date
    response=requests.get(url,headers=headers,params=params).json()
    return response


def fPrintToExcel(ws, dato, antall, prosjekt):
    max_rows = ws.max_row
    ws.cell(column=1,row=max_rows+1, value=dato)
    ws.cell(column=2,row=max_rows+1, value=antall)
    ws.cell(column=3,row=max_rows+1, value=prosjekt)